import win32com.client as win32
import openpyxl
import logging
import time


class ExcelDataUpdations:

    def __init__(self, file_path, visibility=False):
        self.file_path = file_path
        self.xlApp = win32.Dispatch("Excel.Application")
        self.xlApp.Visible = visibility
        self.workbook_win32com = self.xlApp.Workbooks.Open(file_path)
        self.workbook_openpyxl = openpyxl.load_workbook(filename=file_path, data_only=True)

    def get_openpyxl_workbook(self):
        return self.workbook_openpyxl

    def get_win32com_workbook(self):
        return self.workbook_win32com

    def update_external_data(self, wait_time=10):
        """
        Function to update all external link data
        :param wait_time: wait time in seconds before saving file
        :return:
        """
        for con in self.workbook_win32com.connections:
            con.Refresh()
            time.sleep(wait_time)  # move to env file
        self.workbook_win32com.Save()
        return

    def text_to_another_type(self, sheet_name, column, column_no, new_type):
        """
        Function to convert text to new_type
        :param sheet_name: sheet to work on
        :param column: column to work on
        :param column_no: column_no to work on
        :param new_type: new data type to convert
        :return:
        """
        work_sheet = self.workbook_openpyxl[sheet_name]
        worksheet = self.workbook_win32com.Sheets(sheet_name)
        for i in range(2, work_sheet.max_row + 1):
            try:
                worksheet.Cells(i, column_no).Value = eval(new_type)(str(work_sheet[column + str(i)].value).strip())
            except Exception as e:
                try:
                    worksheet.Cells(i, column_no).Value = float(
                        str(worksheet.Cells(i, column_no)).strip())
                except Exception as e:
                    pass
        self.workbook_win32com.Save()
        return

    def filter_replace_value(self, sheet_name, column_no, column, old_value, new_value):
        """
        Function to filter replace value in excel
        :param sheet_name: sheet to work on
        :param column_no: column_no to work on
        :param column: column to work on
        :param old_value: value to replace from
        :param new_value: value to replace to
        :return:
        """
        workbook_win32com = self.xlApp.Workbooks.Open(self.file_path)
        workbook_openpyxl = openpyxl.load_workbook(filename=self.file_path, data_only=True)
        work_sheet = workbook_openpyxl[sheet_name]
        worksheet = workbook_win32com.Sheets(sheet_name)
        for i in range(2, work_sheet.max_row + 1):
            if work_sheet[column + str(i)].value == old_value:
                try:
                    worksheet.Cells(i, column_no).Value = new_value
                except Exception as e:
                    pass
        self.workbook_win32com.Save()
        return

    def update_pivot_tables(self, sheet_name):
        """
        Function to update pivot tables
        :param sheet_name: sheet to work on
        :return:
        """
        worksheet = self.workbook_win32com.Sheets(sheet_name)
        pivotCount = worksheet.PivotTables().Count
        for j in range(1, pivotCount + 1):
            worksheet.PivotTables(j).PivotCache().Refresh()
        self.workbook_win32com.Save()
        return
